import subprocess
import sys

from openpyxl import load_workbook


def test_cli_compare_produces_xlsx(tmp_path):
    a = tmp_path / "a.gcode"
    b = tmp_path / "b.gcode"

    a.write_text(
        """
M83
;Z:0.2
;TYPE:Perimeter
G1 X0 Y0 F6000
G1 X10 Y0 E1.0 F1200
;Z:0.4
;TYPE:Infill
G1 X0 Y0 F6000
G1 X20 Y0 E2.0 F2400
""".lstrip(),
        encoding="utf-8",
    )

    # Different layer heights / Z progression to ensure compare path is exercised.
    b.write_text(
        """
M83
;Z:0.32
;TYPE:Perimeter
G1 X0 Y0 F6000
G1 X10 Y0 E1.2 F1500
;Z:0.64
;TYPE:Infill
G1 X0 Y0 F6000
G1 X20 Y0 E2.4 F2600
""".lstrip(),
        encoding="utf-8",
    )

    script = "gcode_profiler.py"
    # Compare mode writes a combined workbook by default.
    out = tmp_path / "a_vs_b.xlsx"

    subprocess.check_call(
        [sys.executable, script, str(a), "--compare", str(b), "--quiet"],
        cwd=".",
    )

    assert out.exists()
    wb = load_workbook(out)
    assert "Dashboard" in wb.sheetnames
    assert "Layers" in wb.sheetnames
    # Compare sheets should exist when --compare is used.
    assert any(s.startswith("Compare") for s in wb.sheetnames)


def test_cli_compare_with_compare_config(tmp_path):
    a = tmp_path / "a.gcode"
    b = tmp_path / "b.gcode"
    a.write_text(
        """
M83
;Z:0.2
;TYPE:Perimeter
G1 X0 Y0 F6000
G1 X10 Y0 E1.0 F1200
;Z:0.4
;TYPE:Infill
G1 X0 Y0 F6000
G1 X20 Y0 E2.0 F2400
""".lstrip(),
        encoding="utf-8",
    )
    b.write_text(
        """
M83
;Z:0.32
;TYPE:Perimeter
G1 X0 Y0 F6000
G1 X10 Y0 E1.2 F1500
;Z:0.64
;TYPE:Infill
G1 X0 Y0 F6000
G1 X20 Y0 E2.4 F2600
""".lstrip(),
        encoding="utf-8",
    )

    cfg_a = tmp_path / "a.ini"
    cfg_b = tmp_path / "b.ini"
    cfg_a.write_text("filament_max_volumetric_speed = 12\nmax_print_speed = 200\n", encoding="utf-8")
    cfg_b.write_text("filament_max_volumetric_speed = 20\nmax_print_speed = 350\n", encoding="utf-8")

    script = "gcode_profiler.py"
    out = tmp_path / "a_vs_b.xlsx"

    subprocess.check_call(
        [
            sys.executable,
            script,
            str(a),
            "--config",
            str(cfg_a),
            "--compare",
            str(b),
            "--compare-config",
            str(cfg_b),
            "--quiet",
        ],
        cwd=".",
    )

    assert out.exists()
    wb = load_workbook(out)
    assert "Dashboard" in wb.sheetnames
    assert any(s.startswith("Compare") for s in wb.sheetnames)
