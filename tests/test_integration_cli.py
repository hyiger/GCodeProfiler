import subprocess
import sys

from openpyxl import load_workbook


def test_cli_produces_xlsx(tmp_path):
    gcode = tmp_path / "sample.gcode"
    gcode.write_text(
        """
M83
;Z:0.2
;TYPE:Perimeter
G1 X0 Y0 F6000
G1 X10 Y0 E1.0 F1200
;Z:0.4
;TYPE:Infill
G1 X0 Y0 F6000
G1 X20 Y0 E2.0 F2400
""".lstrip(),
        encoding="utf-8",
    )

    script = "gcode_profiler.py"
    out = gcode.with_suffix(".xlsx")

    subprocess.check_call(
        [sys.executable, script, str(gcode), "--quiet"],
        cwd=".",
    )

    assert out.exists()
    wb = load_workbook(out)
    assert "Dashboard" in wb.sheetnames
    assert "Layers" in wb.sheetnames
    ws_layers = wb["Layers"]
    assert ws_layers.max_row >= 2
