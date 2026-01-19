import math
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, ScatterChart, Series
from openpyxl.chart.series_factory import SeriesFactory
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
# RichText support differs across openpyxl versions.
# Some versions (e.g. in certain CI environments) do not expose RichText.
try:
    from openpyxl.drawing.text import RichText, Paragraph, ParagraphProperties, CharacterProperties
except Exception:  # pragma: no cover
    RichText = None
    Paragraph = None
    ParagraphProperties = None
    CharacterProperties = None


from .stats import weighted_quantile, make_bins, bin_counts
from .gcode_parser import filament_area_mm2
from .config_ini import _ini_value_to_float

def set_basic_column_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def write_xlsx(
    moves,
    layer_z_map,
    out_path: str,
    bins: int,
    include_legends: bool,
    per_layer_only: bool,
    top_n_slowest: int,
    filament_diameter_mm: float,
    filament_density_g_cm3: float,
    config_info: dict | None = None,
    layout: str = "compact",
    run_label: str = "A",
    min_peak_segment_time_s: float = 0.05,
    compare_runs: list | None = None,
    top_n_segments: int = 200,
    status_cb=None,
):
    def _status(msg: str):
        if status_cb is not None:
            status_cb(msg)

    _status("Creating workbook")
    wb = Workbook()

    # Geometry for filament usage calculations
    area_mm2 = filament_area_mm2(float(filament_diameter_mm))

    # Dashboard (first sheet)
    ws_dash = wb.active
    ws_dash.title = "Dashboard"

    # Moves sheet
    ws_moves = wb.create_sheet("Moves")

    _status("Populating Moves")

    headers = [
        "layer",
        "z",
        "type",
        "cmd",
        "x0",
        "y0",
        "z0",
        "x1",
        "y1",
        "z1",
        "dist_mm",
        "de_mm",
        "time_s",
        "speed_mm_s",
        "flow_mm3_s",
        "fan_pct",
        "hotend_C",
        "bed_C",
        "chamber_C",
    ]

    ws_moves.append(headers)

    if not per_layer_only:
        for m in moves:
            ws_moves.append([m.get(h) for h in headers])

    for cell in ws_moves["C"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    set_basic_column_widths(
        ws_moves,
        {
            "A": 8,
            "B": 10,
            "C": 30,
            "D": 6,
            "K": 12,
            "L": 12,
            "M": 10,
            "N": 12,
            "O": 12,
            "P": 10,
        },
    )

    # Layers sheet
    ws_layers = wb.create_sheet("Layers")

    _status("Aggregating per-layer metrics")
    ws_layers.append(
        [
            "layer",
            "z_mm",
            "layer_height_mm",
            "time_s",
            "dist_mm",
            "extrusion_mm",
            "avg_speed_mm_s",
            "avg_flow_mm3_s",
            "peak_speed_mm_s",
            "p95_speed_mm_s",
            "p99_speed_mm_s",
            "peak_flow_mm3_s",
            "p95_flow_mm3_s",
            "p99_flow_mm3_s",
            "flow_headroom_p99_mm3_s",
            "speed_headroom_p99_mm_s",
            "travel_time_s",
            "travel_dist_mm",
            "extrude_time_s",
            "retract_count",
            "retract_mm",
            "dynamics_score",
            "over_flow_time_pct",
            "over_speed_time_pct",
            "avg_fan_pct",
            "hotend_set_C",
            "bed_set_C",
            "chamber_set_C",
        ]
    )

    by_layer = defaultdict(list)
    for m in moves:
        by_layer[m["layer"]].append(m)

    layers_sorted = sorted(by_layer.keys())
    prev_z = None

    # Track last known setpoints per layer for cleaner charts
    last_hotend = None
    last_bed = None
    last_chamber = None

    for L in layers_sorted:
        ms = by_layer[L]
        z_val = layer_z_map.get(L, ms[-1]["z"])
        layer_h = (z_val - prev_z) if (prev_z is not None and z_val is not None) else None
        if z_val is not None:
            prev_z = z_val

        t = sum(m["time_s"] for m in ms)
        d = sum(m["dist_mm"] for m in ms)
        e = sum(m["de_mm"] for m in ms)

        # Per-layer worst-case / percentile metrics (tuning-oriented)
        sp_vals = [m["speed_mm_s"] for m in ms if m.get("speed_mm_s") is not None and (m.get("dist_mm") or 0) > 0]
        sp_w = [m["time_s"] for m in ms if m.get("speed_mm_s") is not None and (m.get("dist_mm") or 0) > 0]
        fl_vals = [m["flow_mm3_s"] for m in ms if m.get("flow_mm3_s") is not None and (m.get("flow_mm3_s") or 0) > 0]
        fl_w = [m["time_s"] for m in ms if m.get("flow_mm3_s") is not None and (m.get("flow_mm3_s") or 0) > 0]

        peak_speed = max(sp_vals) if sp_vals else None
        p95_speed = weighted_quantile(sp_vals, sp_w, 0.95) if sp_vals else None
        p99_speed = weighted_quantile(sp_vals, sp_w, 0.99) if sp_vals else None
        peak_flow = max(fl_vals) if fl_vals else None
        p95_flow = weighted_quantile(fl_vals, fl_w, 0.95) if fl_vals else None
        p99_flow = weighted_quantile(fl_vals, fl_w, 0.99) if fl_vals else None

        # Travel / extrusion / retraction diagnostics
        travel_time = sum(m["time_s"] for m in ms if (m.get("de_mm") or 0.0) == 0.0 and (m.get("dist_mm") or 0.0) > 0.0)
        travel_dist = sum(m["dist_mm"] for m in ms if (m.get("de_mm") or 0.0) == 0.0 and (m.get("dist_mm") or 0.0) > 0.0)
        extrude_time = sum(m["time_s"] for m in ms if (m.get("de_mm") or 0.0) > 0.0 and (m.get("time_s") or 0.0) > 0.0)
        retract_moves = [m for m in ms if (m.get("de_mm") or 0.0) < 0.0]
        retract_count = len(retract_moves)
        retract_mm = -sum(m.get("de_mm") or 0.0 for m in retract_moves)

        # Simple dynamics proxy: count short, fast extrusion segments (ringing / PA sensitivity proxy)
        short_fast = 0
        for m in ms:
            if (m.get("de_mm") or 0.0) > 0.0 and (m.get("dist_mm") or 0.0) > 0.0:
                if (m.get("dist_mm") or 0.0) < 0.6 and (m.get("speed_mm_s") or 0.0) > 50.0:
                    short_fast += 1
        dynamics_score = short_fast

        flow_limit = (config_info or {}).get("filament_max_volumetric_speed")
        speed_limit = (config_info or {}).get("max_print_speed")
        flow_headroom = None
        speed_headroom = None
        try:
            if flow_limit is not None and p99_flow is not None:
                flow_headroom = float(flow_limit) - float(p99_flow)
        except Exception:
            pass
        try:
            if speed_limit is not None and p99_speed is not None:
                speed_headroom = float(speed_limit) - float(p99_speed)
        except Exception:
            pass
        over_flow_pct = None
        over_speed_pct = None
        if t and t > 0:
            if flow_limit is not None:
                try:
                    fl_lim = float(flow_limit)
                    over_t = sum(m["time_s"] for m in ms if (m.get("flow_mm3_s") or 0) > fl_lim)
                    over_flow_pct = over_t / t
                except Exception:
                    pass
            if speed_limit is not None:
                try:
                    sp_lim = float(speed_limit)
                    over_t = sum(m["time_s"] for m in ms if (m.get("speed_mm_s") or 0) > sp_lim)
                    over_speed_pct = over_t / t
                except Exception:
                    pass

        if t > 0:
            avg_speed = d / t
            avg_flow = sum(m["flow_mm3_s"] * m["time_s"] for m in ms) / t
            fan_pairs = [(m["fan_pct"], m["time_s"]) for m in ms if m["fan_pct"] is not None]
            avg_fan = (sum(v * w for v, w in fan_pairs) / sum(w for _, w in fan_pairs)) if fan_pairs else None
        else:
            avg_speed = None
            avg_flow = None
            avg_fan = None

        for m in ms:
            if m["hotend_C"] is not None:
                last_hotend = m["hotend_C"]
            if m["bed_C"] is not None:
                last_bed = m["bed_C"]
            if m["chamber_C"] is not None:
                last_chamber = m["chamber_C"]

        ws_layers.append([
            L,
            z_val,
            layer_h,
            t,
            d,
            e,
            avg_speed,
            avg_flow,
            peak_speed,
            p95_speed,
            p99_speed,
            peak_flow,
            p95_flow,
            p99_flow,
            flow_headroom,
            speed_headroom,
            travel_time,
            travel_dist,
            extrude_time,
            retract_count,
            retract_mm,
            dynamics_score,
            over_flow_pct,
            over_speed_pct,
            avg_fan,
            last_hotend,
            last_bed,
            last_chamber,
        ])

    # Optional reference columns from config.ini (used for nicer chart scaling / reference lines)
    if config_info:
        ref_flow = config_info.get('filament_max_volumetric_speed')
        ref_speed = config_info.get('max_print_speed')
        ref_lh_max = config_info.get('max_layer_height')

        # Add columns only if at least one reference value exists
        if any(v is not None for v in (ref_flow, ref_speed, ref_lh_max)):
            base_cols = ws_layers.max_column
            # Headers
            if ref_flow is not None:
                ws_layers.cell(row=1, column=base_cols + 1, value='ref_flow_max_mm3_s')
                for r in range(2, ws_layers.max_row + 1):
                    ws_layers.cell(row=r, column=base_cols + 1, value=float(ref_flow))
                base_cols += 1
            if ref_speed is not None:
                ws_layers.cell(row=1, column=base_cols + 1, value='ref_speed_max_mm_s')
                for r in range(2, ws_layers.max_row + 1):
                    ws_layers.cell(row=r, column=base_cols + 1, value=float(ref_speed))
                base_cols += 1
            if ref_lh_max is not None:
                ws_layers.cell(row=1, column=base_cols + 1, value='ref_layerheight_max_mm')
                for r in range(2, ws_layers.max_row + 1):
                    ws_layers.cell(row=r, column=base_cols + 1, value=float(ref_lh_max))
                base_cols += 1

    set_basic_column_widths(
        ws_layers,
        {
            "A": 8,
            "B": 10,
            "C": 16,
            "D": 12,
            "E": 12,
            "F": 14,
            "G": 16,
            "H": 16,
            "I": 16,
            "J": 16,
            "K": 16,
            "L": 16,
            "M": 16,
            "N": 16,
            "O": 18,
            "P": 18,
            "Q": 14,
            "R": 14,
            "S": 14,
            "T": 14,
            "U": 14,
            "V": 14,
            "W": 14,
            "X": 14,
            "Y": 12,
            "Z": 12,
            "AA": 12,
            "AB": 12,
        },
    )

    # Nice-to-have: conditional formatting for outliers (useful at-a-glance).
    # These are driven by --config (or explicit CLI overrides) when available.
    if config_info:
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        max_flow = config_info.get('filament_max_volumetric_speed')
        max_speed = config_info.get('max_print_speed')
        max_lh = config_info.get('max_layer_height')
        min_lh = config_info.get('min_layer_height')

        last = ws_layers.max_row
        # Avg speed (col G)
        if max_speed is not None:
            ws_layers.conditional_formatting.add(
                f"G2:G{last}",
                CellIsRule(operator='greaterThan', formula=[str(float(max_speed))], fill=yellow_fill)
            )
            # Peak/P95/P99 speed
            ws_layers.conditional_formatting.add(
                f"I2:I{last}",
                CellIsRule(operator='greaterThan', formula=[str(float(max_speed))], fill=yellow_fill)
            )
            ws_layers.conditional_formatting.add(
                f"J2:J{last}",
                CellIsRule(operator='greaterThan', formula=[str(float(max_speed))], fill=yellow_fill)
            )
            ws_layers.conditional_formatting.add(
                f"K2:K{last}",
                CellIsRule(operator='greaterThan', formula=[str(float(max_speed))], fill=yellow_fill)
            )
        # Avg flow (col H)
        if max_flow is not None:
            ws_layers.conditional_formatting.add(
                f"H2:H{last}",
                CellIsRule(operator='greaterThan', formula=[str(float(max_flow))], fill=red_fill)
            )
            # Peak/P95/P99 flow
            ws_layers.conditional_formatting.add(
                f"L2:L{last}",
                CellIsRule(operator='greaterThan', formula=[str(float(max_flow))], fill=red_fill)
            )
            ws_layers.conditional_formatting.add(
                f"M2:M{last}",
                CellIsRule(operator='greaterThan', formula=[str(float(max_flow))], fill=red_fill)
            )
            ws_layers.conditional_formatting.add(
                f"N2:N{last}",
                CellIsRule(operator='greaterThan', formula=[str(float(max_flow))], fill=red_fill)
            )
        # Layer height bounds (col C)
        if max_lh is not None:
            ws_layers.conditional_formatting.add(
                f"C2:C{last}",
                CellIsRule(operator='greaterThan', formula=[str(float(max_lh))], fill=yellow_fill)
            )
        if min_lh is not None:
            ws_layers.conditional_formatting.add(
                f"C2:C{last}",
                CellIsRule(operator='lessThan', formula=[str(float(min_lh))], fill=yellow_fill)
            )

    # Legends
    def add_legend_sheet(name, values, unit_label, forced_min=None, forced_max=None):
        ws = wb.create_sheet(name)
        clean = [v for v in values if v is not None]
        if not clean:
            ws.append(["No data"])
            return ws
        vmin, vmax = min(clean), max(clean)
        if forced_min is not None:
            try:
                vmin = float(forced_min)
            except Exception:
                pass
        if forced_max is not None:
            try:
                vmax = float(forced_max)
            except Exception:
                pass
        if vmax < vmin:
            vmin, vmax = vmax, vmin
        ws.append(["min", vmin])
        ws.append(["max", vmax])
        ws.append([])
        ws.append(["bin", f"range ({unit_label})", "count"])
        bins_spec = make_bins(vmin, vmax, bins)
        counts = bin_counts(clean, bins_spec)
        for i, ((lo, hi), c) in enumerate(zip(bins_spec, counts), start=1):
            ws.append([i, f"{lo:.6g} – {hi:.6g}", c])
        set_basic_column_widths(ws, {"A": 8, "B": 24, "C": 10})
        return ws

    _status("Building legend sheets")
    if include_legends:
        speeds = [m["speed_mm_s"] for m in moves if m["speed_mm_s"] is not None and m["dist_mm"] > 0]
        flows = [m["flow_mm3_s"] for m in moves if m["flow_mm3_s"] is not None and m["flow_mm3_s"] > 0]
        fans = [m["fan_pct"] for m in moves if m["fan_pct"] is not None]
        hotends = [m["hotend_C"] for m in moves if m["hotend_C"] is not None]
        beds = [m["bed_C"] for m in moves if m["bed_C"] is not None]

        layer_heights = []
        for row in ws_layers.iter_rows(min_row=2, values_only=True):
            lh = row[2]
            if lh is not None and lh > 0:
                layer_heights.append(lh)

        add_legend_sheet("Legend_Speed", speeds, "mm/s", forced_min=0, forced_max=(config_info or {}).get("max_print_speed"))
        add_legend_sheet("Legend_Flow_mm3s", flows, "mm³/s", forced_min=0, forced_max=(config_info or {}).get("filament_max_volumetric_speed"))
        add_legend_sheet("Legend_Fan_pct", fans, "%")
        add_legend_sheet("Legend_Temp_C", hotends, "°C")
        add_legend_sheet("Legend_Bed_C", beds, "°C")
        add_legend_sheet("Legend_LayerHeight_mm", layer_heights, "mm", forced_min=(config_info or {}).get("min_layer_height"), forced_max=(config_info or {}).get("max_layer_height"))
        ws_ft = wb.create_sheet("Legend_FeatureType")
        c = Counter(m["type"] for m in moves if m.get("type"))

        # Totals for percentages + filament usage
        total_time_s = sum(m.get("time_s", 0.0) or 0.0 for m in moves)
        total_de_mm = sum(m.get("de_mm", 0.0) or 0.0 for m in moves if (m.get("de_mm", 0.0) or 0.0) > 0)

        # Excel stores time as days. We'll store time as days and format as [h]:mm:ss
        ws_ft.append(["Feature type", "Time", "Percentage", "Used filament (m)", "Used filament (g)", "Move count"])

        for t, n in c.most_common():
            ms = [m for m in moves if m.get("type") == t]
            time_s = sum(m.get("time_s", 0.0) or 0.0 for m in ms)
            de_mm = sum((m.get("de_mm", 0.0) or 0.0) for m in ms if (m.get("de_mm", 0.0) or 0.0) > 0)

            # Percentage of total time
            pct = (time_s / total_time_s) if total_time_s > 0 else 0.0

            used_m = de_mm / 1000.0

            # grams = volume_cm3 * density_g_cm3
            vol_mm3 = de_mm * area_mm2
            vol_cm3 = vol_mm3 / 1000.0
            used_g = vol_cm3 * float(filament_density_g_cm3)

            ws_ft.append([
                t,
                time_s / 86400.0,
                pct,
                used_m,
                used_g,
                n,
            ])

        # Formatting
        for cell in ws_ft["A"]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for cell in ws_ft["B"][1:]:
            cell.number_format = "[h]:mm:ss"
        for cell in ws_ft["C"][1:]:
            cell.number_format = "0.0%"
        for cell in ws_ft["D"][1:]:
            cell.number_format = "0.00"
        for cell in ws_ft["E"][1:]:
            cell.number_format = "0.00"

        set_basic_column_widths(ws_ft, {"A": 34, "B": 12, "C": 12, "D": 16, "E": 16, "F": 12})

        # Feature-type flow/speed limits summary (tuning-focused)
        ws_ff = wb.create_sheet("FeatureType_Flow")
        ws_ff.append([
            "Feature type",
            "Time",
            "Time %",
            "Used filament (m)",
            "Used filament (g)",
            "Peak speed (mm/s)",
            "P95 speed (mm/s)",
            "Peak flow (mm³/s)",
            "P95 flow (mm³/s)",
            "Over flow limit % time",
            "Over speed limit % time",
            "Move count",
        ])

        flow_limit = (config_info or {}).get("filament_max_volumetric_speed")
        speed_limit = (config_info or {}).get("max_print_speed")
        try:
            flow_limit_f = float(flow_limit) if flow_limit is not None else None
        except Exception:
            flow_limit_f = None
        try:
            speed_limit_f = float(speed_limit) if speed_limit is not None else None
        except Exception:
            speed_limit_f = None

        for t, n in c.most_common():
            ms = [m for m in moves if m.get("type") == t]
            time_s = sum(m.get("time_s", 0.0) or 0.0 for m in ms)
            de_mm = sum((m.get("de_mm", 0.0) or 0.0) for m in ms if (m.get("de_mm", 0.0) or 0.0) > 0)
            pct = (time_s / total_time_s) if total_time_s > 0 else 0.0
            used_m = de_mm / 1000.0
            vol_cm3 = (de_mm * area_mm2) / 1000.0
            used_g = vol_cm3 * float(filament_density_g_cm3)

            sp_vals = [m["speed_mm_s"] for m in ms if m.get("speed_mm_s") is not None and (m.get("dist_mm") or 0) > 0]
            sp_w = [m["time_s"] for m in ms if m.get("speed_mm_s") is not None and (m.get("dist_mm") or 0) > 0]
            fl_vals = [m["flow_mm3_s"] for m in ms if m.get("flow_mm3_s") is not None and (m.get("flow_mm3_s") or 0) > 0]
            fl_w = [m["time_s"] for m in ms if m.get("flow_mm3_s") is not None and (m.get("flow_mm3_s") or 0) > 0]
            p95_speed = weighted_quantile(sp_vals, sp_w, 0.95) if sp_vals else None
            p99_speed = weighted_quantile(sp_vals, sp_w, 0.99) if sp_vals else None
            peak_speed_raw = max(sp_vals) if sp_vals else None
            # Spike suppression: cap extreme peaks to a high percentile when they look like single-segment noise.
            if peak_speed_raw is not None and p99_speed is not None and peak_speed_raw > 1.5 * p99_speed:
                peak_speed = p99_speed
            else:
                peak_speed = peak_speed_raw

            p95_flow = weighted_quantile(fl_vals, fl_w, 0.95) if fl_vals else None
            p99_flow = weighted_quantile(fl_vals, fl_w, 0.99) if fl_vals else None
            peak_flow_raw = max(fl_vals) if fl_vals else None
            if peak_flow_raw is not None and p99_flow is not None and peak_flow_raw > 1.5 * p99_flow:
                peak_flow = p99_flow
            else:
                peak_flow = peak_flow_raw

            over_flow_pct = None
            over_speed_pct = None
            if time_s and time_s > 0:
                if flow_limit_f is not None:
                    over_t = sum(m.get("time_s", 0.0) or 0.0 for m in ms if (m.get("flow_mm3_s") or 0) > flow_limit_f)
                    over_flow_pct = over_t / time_s
                if speed_limit_f is not None:
                    over_t = sum(m.get("time_s", 0.0) or 0.0 for m in ms if (m.get("speed_mm_s") or 0) > speed_limit_f)
                    over_speed_pct = over_t / time_s

            ws_ff.append([
                t,
                time_s / 86400.0,
                pct,
                used_m,
                used_g,
                peak_speed,
                p95_speed,
                peak_flow,
                p95_flow,
                over_flow_pct,
                over_speed_pct,
                n,
            ])

        for cell in ws_ff["A"]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for cell in ws_ff["B"][1:]:
            cell.number_format = "[h]:mm:ss"
        for col in ("C", "J", "K"):
            for cell in ws_ff[col][1:]:
                cell.number_format = "0.0%"
        for col in ("D", "E"):
            for cell in ws_ff[col][1:]:
                cell.number_format = "0.00"
        for col in ("F", "G", "H", "I"):
            for cell in ws_ff[col][1:]:
                cell.number_format = "0.00"

        set_basic_column_widths(ws_ff, {"A": 34, "B": 12, "C": 10, "D": 16, "E": 16, "F": 16, "G": 16, "H": 16, "I": 16, "J": 18, "K": 18, "L": 12})

    # Top segments by volumetric flow (helps find brief spikes)
    _status("Computing top flow segments")
    ws_top = wb.create_sheet("Top_Flow_Segments")
    ws_top.append([
        "rank", "layer", "type", "z_mm", "x0", "y0", "x1", "y1",
        "dist_mm", "de_mm", "time_s", "speed_mm_s", "flow_mm3_s"
    ])
    extrude_segs = [m for m in moves if (m.get("de_mm") or 0.0) > 0.0 and (m.get("time_s") or 0.0) > 0.0 and (m.get("flow_mm3_s") or 0.0) > 0.0]
    extrude_segs.sort(key=lambda m: (m.get("flow_mm3_s") or 0.0), reverse=True)
    for idx, m in enumerate(extrude_segs[:max(1, int(top_n_segments))], start=1):
        ws_top.append([
            idx,
            m.get("layer"),
            m.get("type"),
            m.get("z"),
            m.get("x0"), m.get("y0"), m.get("x1"), m.get("y1"),
            m.get("dist_mm"), m.get("de_mm"), m.get("time_s"),
            m.get("speed_mm_s"), m.get("flow_mm3_s"),
        ])
    set_basic_column_widths(ws_top, {"A": 6, "B": 8, "C": 28, "D": 10, "E": 10, "F": 10, "G": 10, "H": 10, "I": 10, "J": 10, "K": 10, "L": 12, "M": 12})

    # Flow histogram by feature type (time-weighted) for quick diagnosis
    _status("Computing feature-type flow histograms")
    ws_fh = wb.create_sheet("FeatureFlow_Hist")
    ws_fh.append(["Feature type", "bin_lo", "bin_hi", "time_s", "time_pct"])
    flow_values = [m.get("flow_mm3_s") for m in extrude_segs]
    if flow_values:
        lo = 0.0
        hi = max(flow_values)
        if (config_info or {}).get("filament_max_volumetric_speed") is not None:
            try:
                hi = max(hi, float((config_info or {}).get("filament_max_volumetric_speed")))
            except Exception:
                pass
        bins_spec = make_bins(lo, hi, bins)
        total_time = sum(m.get("time_s") or 0.0 for m in extrude_segs)
        by_type = defaultdict(list)
        for m in extrude_segs:
            by_type[m.get("type") or "UNKNOWN"].append(m)
        for t, ms in sorted(by_type.items(), key=lambda kv: sum(m.get("time_s") or 0.0 for m in kv[1]), reverse=True):
            for (b_lo, b_hi) in bins_spec:
                bt = 0.0
                for m in ms:
                    v = m.get("flow_mm3_s") or 0.0
                    if b_lo <= v < b_hi or (b_hi == bins_spec[-1][1] and b_lo <= v <= b_hi):
                        bt += m.get("time_s") or 0.0
                pct = (bt / total_time) if total_time > 0 else None
                ws_fh.append([t, b_lo, b_hi, bt, pct])
    set_basic_column_widths(ws_fh, {"A": 34, "B": 12, "C": 12, "D": 12, "E": 12})
    for cell in ws_fh["E"][1:]:
        cell.number_format = "0.0%"

    # Top N slowest layers
    ws_top = wb.create_sheet("Top_Slowest_Layers")
    ws_top.append(["rank", "layer", "time_s", "z_mm", "avg_speed_mm_s", "avg_flow_mm3_s", "avg_fan_pct", "hotend_set_C", "bed_set_C", "chamber_set_C"])

    layer_rows = []
    for r in ws_layers.iter_rows(min_row=2, values_only=True):
        # row: layer, z, lh, time, dist, extrusion, avg_speed, avg_flow, peak_speed, p95_speed,
        #      peak_flow, p95_flow, over_flow_pct, over_speed_pct, avg_fan, hotend, bed, chamber
        layer_rows.append(r)

    layer_rows_sorted = sorted(layer_rows, key=lambda r: (r[3] if r[3] is not None else -1), reverse=True)
    top_n = max(1, int(top_n_slowest))
    for i, r in enumerate(layer_rows_sorted[:top_n], start=1):
        ws_top.append([i, r[0], r[3], r[1], r[6], r[7], r[24], r[25], r[26], r[27]])

    set_basic_column_widths(ws_top, {"A": 6, "B": 8, "C": 12, "D": 10, "E": 16, "F": 16, "G": 12, "H": 14, "I": 12, "J": 12})

    _status("Building Dashboard")
    # Dashboard charts (on first sheet)
    # Excel column letters go A..Z, AA.. etc. Using chr() past 'Z' becomes '[' and breaks openpyxl.
    # Make enough columns available so we can anchor a true two-column dashboard
    # with plenty of horizontal separation (Excel chart bounding boxes can be wider
    # than expected on some platforms/zoom levels).
    ws_dash["A1"] = "Dashboard (charts are generated from Layers / Legend_* / Top_Slowest_Layers)."

    # Config summary (optional)
    if config_info:
        ws_dash["B1"] = "Config summary"
        pairs = [
            ("Nozzle (mm)", config_info.get("nozzle_diameter")),
            ("Filament (mm)", config_info.get("filament_diameter")),
            ("Density (g/cm³)", config_info.get("filament_density")),
            ("Max volumetric (mm³/s)", config_info.get("filament_max_volumetric_speed")),
            ("Max print speed (mm/s)", config_info.get("max_print_speed")),
            ("Layer height (mm)", config_info.get("layer_height")),
            ("First layer (mm)", config_info.get("first_layer_height")),
        ]
        row = 2
        for k, v in pairs:
            if v is None:
                continue
            ws_dash[f"G{row}"] = k
            ws_dash[f"H{row}"] = float(v)
            row += 1
        set_basic_column_widths(ws_dash, {"G": 22, "H": 18})

    # Feature type legend table (similar to PrusaSlicer UI)
    if include_legends and "Legend_FeatureType" in wb.sheetnames:
        ws_dash["B2"] = "Feature type"
        ws_dash["C2"] = "Time"
        ws_dash["D2"] = "Percentage"
        ws_dash["E2"] = "Used filament (m)"
        ws_dash["F2"] = "Used filament (g)"

        # Copy top feature types (by time) from Legend_FeatureType
        ws_ft = wb["Legend_FeatureType"]
        # Data starts at row 2
        # Sort by time (column B) descending
        rows = []
        for r in ws_ft.iter_rows(min_row=2, values_only=True):
            rows.append(r)
        rows.sort(key=lambda r: (r[1] if r and r[1] is not None else 0), reverse=True)
        max_rows = min(10, len(rows))
        for i in range(max_rows):
            r = rows[i]
            out_row = 3 + i
            ws_dash[f"B{out_row}"] = r[0]
            ws_dash[f"C{out_row}"] = r[1]
            ws_dash[f"D{out_row}"] = r[2]
            ws_dash[f"E{out_row}"] = r[3]
            ws_dash[f"F{out_row}"] = r[4]
            ws_dash[f"C{out_row}"].number_format = "[h]:mm:ss"
            ws_dash[f"D{out_row}"].number_format = "0.0%"
            ws_dash[f"E{out_row}"].number_format = "0.00"
            ws_dash[f"F{out_row}"].number_format = "0.00"

        set_basic_column_widths(ws_dash, {"A": 3, "B": 28, "C": 12, "D": 12, "E": 18, "F": 18})


    max_layer_row = ws_layers.max_row

    # Locate optional reference columns in Layers (if config was provided)
    ref_cols = {}
    header_row = [c.value for c in ws_layers[1]]
    for idx, name in enumerate(header_row, start=1):
        if name in ("ref_flow_max_mm3_s", "ref_speed_max_mm_s", "ref_layerheight_max_mm"):
            ref_cols[name] = idx
    cats_layers = Reference(ws_layers, min_col=1, min_row=2, max_row=max_layer_row)

    # NOTE: openpyxl chart sizes are in "Excel" units (roughly inches).
    # Keep charts modest and place them with generous spacing to avoid overlap.
    # NOTE: Chart object sizes are in "Excel units" (roughly inches). Excel's
    # rendered bounding boxes can be larger than you'd expect, so we keep charts
    # relatively small and leave generous horizontal/vertical gaps.
    # Reduce label clutter: skip most x labels and rotate the remainder.
    # Keep a floor of 1 (show everything) for short prints.
    # We'll set the skip factor after deciding the dashboard layout.
    label_skip = 1

    def _axis_font(size_pt: float):
        """Create a RichText object that sets axis tick label font size.

        openpyxl doesn't expose a simple `font` property for axis tick labels.
        Setting `axis.txPr` with a RichText paragraph is the most reliable
        cross-platform way to control tick label size.
        """
        if RichText is None or Paragraph is None or ParagraphProperties is None or CharacterProperties is None:
            return None
        # Excel stores font size in 1/100 pt.
        sz = int(round(size_pt * 100))
        ppr = ParagraphProperties(defRPr=CharacterProperties(sz=sz))
        return RichText(p=[Paragraph(pPr=ppr)])

    def _style_axis(axis):
        # Force axis + tick labels to be shown.
        # Excel can sometimes hide tick labels if an axis is considered unused.
        axis.delete = False
        axis.tickLblPos = "nextTo"
        # Tick marks help readability.
        axis.majorTickMark = "out"
        axis.minorTickMark = "none"
        # Ensure tick label font is readable and doesn't balloon.
        # 10pt is a good default for Y-axis.
        try:
            _rt = _axis_font(10)
            if _rt is not None:
                axis.txPr = _rt
        except Exception:
            pass
        return axis

    def _style_x_axis(axis):
        _style_axis(axis)
        # Keep x-axis labels readable and consistent with y-axis.
        #
        # For category axes (older LineChart usage), tickLblSkip reduces clutter.
        # For numeric axes (ScatterChart), tickLblSkip is ignored by Excel.
        try:
            axis.tickLblSkip = label_skip
            axis.tickMarkSkip = label_skip
        except Exception:
            pass
        # Avoid rotated, oversized labels (these tend to collide).
        axis.textRotation = 0
        # Match y-axis tick label size.
        try:
            _rt = _axis_font(10)
            if _rt is not None:
                axis.txPr = _rt
        except Exception:
            pass
        return axis

    def add_line_chart(title, y_title, min_col, anchor, width=13, height=7, max_col=None, extra_series_cols=None):
        """Add a per-layer trend chart.

        We intentionally use a ScatterChart (numeric X axis) instead of a LineChart
        (category axis). Category axes become unreadable for prints with hundreds
        of layers because Excel tries to render too many tick labels and may scale
        the font up aggressively. ScatterChart gives us a numeric axis where we can
        control the tick interval (majorUnit) and keep labels readable.
        """
        ch = ScatterChart()
        ch.title = title
        ch.y_axis.title = y_title
        ch.x_axis.title = "layer"

        # Axes styling
        _style_axis(ch.y_axis)
        _style_axis(ch.x_axis)
        ch.legend = None

        # Numeric X values (layer index)
        xvalues = Reference(ws_layers, min_col=1, min_row=2, max_row=max_layer_row)

        def _add_series(col_idx: int):
            # Use header cell as series title.
            yvalues = Reference(ws_layers, min_col=col_idx, min_row=2, max_row=max_layer_row)
            # Use openpyxl's SeriesFactory for cross-version compatibility.
            s = SeriesFactory(yvalues, xvalues=xvalues, title=None, title_from_data=False)
            try:
                s.title = ws_layers.cell(row=1, column=col_idx).value
            except Exception:
                pass
            ch.series.append(s)

        # Primary series columns
        if max_col is None:
            _add_series(int(min_col))
        else:
            for c in range(int(min_col), int(max_col) + 1):
                _add_series(c)

        # Optional additional series (e.g. config reference lines)
        if extra_series_cols:
            for col in extra_series_cols:
                if col is None:
                    continue
                try:
                    _add_series(int(col))
                except Exception:
                    pass

        # Tick spacing: keep labels from colliding even when Excel chooses a
        # very large default font for axis labels. We intentionally target a
        # small number of labels.
        try:
            n = max(1, max_layer_row - 1)
            target = 6 if (layout or "compact").strip().lower() == "compact" else 8
            major = max(1, int(math.ceil(n / float(target))))
            ch.x_axis.majorUnit = major
        except Exception:
            pass

        # Make the axis labels smaller and avoid rotation; numeric axis labels are sparse.
        ch.x_axis.textRotation = 0

        ch.height = height
        ch.width = width
        ws_dash.add_chart(ch, anchor)
        return ch

    # Layout: dashboard grid.
    # Excel positions charts in pixel space. To keep the layout stable across
    # platforms/zoom levels we:
    #   - use normal column widths (so anchors correspond to real spacing)
    #   - reserve a left margin so Y-axis tick labels don't get clipped
    #   - support two layouts: compact (default) and wide
    layers_count = max(0, max_layer_row - 1)
    layout = (layout or "compact").strip().lower()
    if layout not in ("compact", "wide"):
        layout = "compact"

    # X-axis label downsampling: keep charts readable when there are hundreds of layers.
    # Target fewer labels in compact layout, more in wide layout.
    target_labels = 8 if layout == "compact" else 12
    label_skip = max(1, layers_count // max(1, target_labels))

    # Column geometry: tighter for compact, more spacious for wide.
    base_col_w = 9 if layout == "compact" else 12
    set_basic_column_widths(ws_dash, {get_column_letter(i): base_col_w for i in range(1, 80)})
    ws_dash.column_dimensions["A"].width = 5  # left margin for y-axis labels

    LEFT = "B"
    RIGHT = "N" if layout == "compact" else "Q"

    # Auto-scale chart width by number of layers.
    # More layers => wider charts for readability (without causing overlap).
    # Clamp to a safe range.
    if layout == "wide":
        CH_W = max(16, min(22, 16 + (layers_count / 120.0)))
        CH_H = 7.6
    else:
        CH_W = max(14, min(18, 14 + (layers_count / 200.0)))
        CH_H = 7.2

    # Vertical spacing: reduce empty gap while keeping charts from touching.
    # Leave space at the top for the Feature Type table.
    R1, R2, R3, R4, R5, R6, R7, R8, R9 = 16, 34, 52, 70, 88, 106, 126, 144, 162

    # Row 1
    time_ch = add_line_chart("Time per Layer (s)", "seconds", 4, f"{LEFT}{R1}", width=CH_W, height=CH_H)
    speed_ch = add_line_chart("Average Speed per Layer (mm/s)", "mm/s", 7, f"{RIGHT}{R1}", width=CH_W, height=CH_H, extra_series_cols=[ref_cols.get("ref_speed_max_mm_s")] if ref_cols.get("ref_speed_max_mm_s") else None)
    if ref_cols.get("ref_speed_max_mm_s"):
        try:
            speed_ch.y_axis.scaling.max = float(config_info.get("max_print_speed")) * 1.1
        except Exception:
            pass

    # Row 2
    flow_ch = add_line_chart("Average Volumetric Flow per Layer (mm³/s)", "mm³/s", 8, f"{LEFT}{R2}", width=CH_W, height=CH_H, extra_series_cols=[ref_cols.get("ref_flow_max_mm3_s")] if ref_cols.get("ref_flow_max_mm3_s") else None)
    if ref_cols.get("ref_flow_max_mm3_s"):
        try:
            flow_ch.y_axis.scaling.max = float(config_info.get("filament_max_volumetric_speed")) * 1.1
        except Exception:
            pass

    # Layer height (column)
    lh_bar = BarChart()
    lh_bar.type = "col"
    lh_bar.title = "Layer Height per Layer (mm)"
    lh_bar.y_axis.title = "mm"
    # Keep layer height chart scale tight using config.ini max_layer_height when available
    cfg_layer_h_max = _ini_value_to_float((config_info or {}).get("max_layer_height"))
    if cfg_layer_h_max is not None:
        try:
            lh_bar.y_axis.scaling.max = float(cfg_layer_h_max) * 1.1
        except Exception:
            pass
    if ref_cols.get("ref_layerheight_max_mm") and config_info and config_info.get("max_layer_height") is not None:
        try:
            lh_bar.y_axis.scaling.max = float(config_info.get("max_layer_height")) * 1.1
        except Exception:
            pass
    lh_bar.x_axis.title = "layer"
    _style_axis(lh_bar.y_axis)
    _style_x_axis(lh_bar.x_axis)
    lh_bar.legend = None
    lh_data = Reference(ws_layers, min_col=3, min_row=1, max_row=max_layer_row)
    lh_bar.add_data(lh_data, titles_from_data=True)
    lh_bar.set_categories(cats_layers)
    lh_bar.height = CH_H
    lh_bar.width = CH_W
    ws_dash.add_chart(lh_bar, f"{RIGHT}{R2}")

    # Row 3
    add_line_chart("Extrusion per Layer (mm of filament)", "mm", 6, f"{LEFT}{R3}", width=CH_W, height=CH_H)
    add_line_chart("Average Fan per Layer (%)", "%", 25, f"{RIGHT}{R3}", width=CH_W, height=CH_H)

    # Row 4
    add_line_chart("Set Temperatures per Layer (°C)", "°C", 26, f"{LEFT}{R4}", width=CH_W, height=CH_H, max_col=28)

    # Histograms: speed + flow (from legends)
    def add_histogram(legend_sheet_name, title, anchor):
        if not include_legends or legend_sheet_name not in wb.sheetnames:
            return
        ws_leg = wb[legend_sheet_name]
        if ws_leg.max_row < 5:
            return
        bar = BarChart()
        bar.type = "col"
        bar.title = title
        bar.y_axis.title = "count"
        bar.x_axis.title = "bin"
        _style_axis(bar.y_axis)
        _style_x_axis(bar.x_axis)
        bar.legend = None
        cats = Reference(ws_leg, min_col=1, min_row=5, max_row=ws_leg.max_row)
        data = Reference(ws_leg, min_col=3, min_row=4, max_row=ws_leg.max_row)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        # Match dashboard sizing so the histograms align with other charts.
        bar.height = CH_H
        bar.width = CH_W
        ws_dash.add_chart(bar, anchor)

    # Row 4 (right): speed histogram
    add_histogram("Legend_Speed", "Speed Bin Counts", f"{RIGHT}{R4}")

    # Row 5: flow histogram + pie
    add_histogram("Legend_Flow_mm3s", "Flow (mm³/s) Bin Counts", f"{LEFT}{R5}")

    # Feature type time share pie (always on dashboard)
    if include_legends and "Legend_FeatureType" in wb.sheetnames:
        ws_ft = wb["Legend_FeatureType"]
        if ws_ft.max_row >= 3:
            pie = PieChart()
            pie.title = "Feature Type Time Share"
            labels = Reference(ws_ft, min_col=1, min_row=2, max_row=ws_ft.max_row)
            data = Reference(ws_ft, min_col=3, min_row=1, max_row=ws_ft.max_row)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.dataLabels = DataLabelList()
            pie.dataLabels.showPercent = True
            pie.height = CH_H
            pie.width = CH_W
            ws_dash.add_chart(pie, f"{RIGHT}{R5}")

    # Bottom row: Top N slowest layers (span both columns)
    if ws_top.max_row >= 3:
        slow_bar = BarChart()
        slow_bar.type = "col"
        slow_bar.title = f"Top {top_n} Slowest Layers (time_s)"
        slow_bar.y_axis.title = "seconds"
        slow_bar.x_axis.title = "layer"
        _style_axis(slow_bar.y_axis)
        _style_x_axis(slow_bar.x_axis)
        slow_bar.legend = None
        slow_data = Reference(ws_top, min_col=3, min_row=1, max_row=ws_top.max_row)
        slow_cats = Reference(ws_top, min_col=2, min_row=2, max_row=ws_top.max_row)
        slow_bar.add_data(slow_data, titles_from_data=True)
        slow_bar.set_categories(slow_cats)
        slow_bar.height = 8
        slow_bar.width = 31
        ws_dash.add_chart(slow_bar, f"{LEFT}{R6}")

    # Tuning-focused: worst-case / percentile charts (keep existing averages too)
    # Columns: I peak_speed, J p95_speed, K p99_speed, L peak_flow, M p95_flow, N p99_flow
    peak_sp = add_line_chart("Peak Speed per Layer (mm/s)", "mm/s", 9, f"{LEFT}{R7}", width=CH_W, height=CH_H,
                             extra_series_cols=[ref_cols.get("ref_speed_max_mm_s")] if ref_cols.get("ref_speed_max_mm_s") else None)
    p95_sp = add_line_chart("P95 Speed per Layer (mm/s)", "mm/s", 10, f"{RIGHT}{R7}", width=CH_W, height=CH_H,
                            extra_series_cols=[ref_cols.get("ref_speed_max_mm_s")] if ref_cols.get("ref_speed_max_mm_s") else None)
    p99_sp = add_line_chart("P99 Speed per Layer (mm/s)", "mm/s", 11, f"{LEFT}{R8}", width=CH_W, height=CH_H,
                            extra_series_cols=[ref_cols.get("ref_speed_max_mm_s")] if ref_cols.get("ref_speed_max_mm_s") else None)
    peak_fl = add_line_chart("Peak Volumetric Flow per Layer (mm³/s)", "mm³/s", 12, f"{RIGHT}{R8}", width=CH_W, height=CH_H,
                             extra_series_cols=[ref_cols.get("ref_flow_max_mm3_s")] if ref_cols.get("ref_flow_max_mm3_s") else None)
    p95_fl = add_line_chart("P95 Volumetric Flow per Layer (mm³/s)", "mm³/s", 13, f"{LEFT}{R9}", width=CH_W, height=CH_H,
                            extra_series_cols=[ref_cols.get("ref_flow_max_mm3_s")] if ref_cols.get("ref_flow_max_mm3_s") else None)
    p99_fl = add_line_chart("P99 Volumetric Flow per Layer (mm³/s)", "mm³/s", 14, f"{RIGHT}{R9}", width=CH_W, height=CH_H,
                            extra_series_cols=[ref_cols.get("ref_flow_max_mm3_s")] if ref_cols.get("ref_flow_max_mm3_s") else None)

    # Scale maxima based on config where available
    if config_info:
        try:
            if config_info.get("max_print_speed") is not None:
                m = float(config_info.get("max_print_speed")) * 1.1
                peak_sp.y_axis.scaling.max = m
                p95_sp.y_axis.scaling.max = m
                p99_sp.y_axis.scaling.max = m
        except Exception:
            pass
        try:
            if config_info.get("filament_max_volumetric_speed") is not None:
                m = float(config_info.get("filament_max_volumetric_speed")) * 1.1
                peak_fl.y_axis.scaling.max = m
                p95_fl.y_axis.scaling.max = m
                p99_fl.y_axis.scaling.max = m
        except Exception:
            pass

    # Compare mode (experimentation)
    # If multiple compares are provided, we build a summary across all.
    # For charts, we align runs on *Z height* (numeric axis) so prints with different layer heights still line up.
    if compare_runs:
        ws_cb = wb.create_sheet("Compare_Layers")

        def _layer_stats_series(moves_x, layer_z_x):
            """Return per-layer stats as a list of dicts sorted by Z.

            Keys: z, layer, time_s, peak_flow, p95_flow, peak_speed, p95_speed
            """
            by = defaultdict(list)
            for m in moves_x:
                by[m["layer"]].append(m)

            rows = []
            for Lx, msx in by.items():
                z = layer_z_x.get(Lx)
                if z is None:
                    z = msx[-1].get("z")

                t = sum(m.get("time_s", 0.0) or 0.0 for m in msx)

                sp_vals = [m["speed_mm_s"] for m in msx if m.get("speed_mm_s") is not None and (m.get("dist_mm") or 0) > 0]
                sp_w = [m.get("time_s") or 0.0 for m in msx if m.get("speed_mm_s") is not None and (m.get("dist_mm") or 0) > 0]

                fl_vals = [m["flow_mm3_s"] for m in msx if (m.get("flow_mm3_s") or 0.0) > 0.0]
                fl_w = [m.get("time_s") or 0.0 for m in msx if (m.get("flow_mm3_s") or 0.0) > 0.0]

                # Filtered peaks to reduce single-move spike noise: ignore segments shorter than min_peak_segment_time_s
                sp_vals_f = [m["speed_mm_s"] for m in msx if m.get("speed_mm_s") is not None and (m.get("dist_mm") or 0) > 0 and (m.get("time_s") or 0.0) >= min_peak_segment_time_s]
                fl_vals_f = [m["flow_mm3_s"] for m in msx if (m.get("flow_mm3_s") or 0.0) > 0.0 and (m.get("time_s") or 0.0) >= min_peak_segment_time_s]

                peak_speed = max(sp_vals_f) if sp_vals_f else (max(sp_vals) if sp_vals else None)
                p95_speed = weighted_quantile(sp_vals, sp_w, 0.95) if sp_vals else None
                peak_flow = max(fl_vals_f) if fl_vals_f else (max(fl_vals) if fl_vals else None)
                p95_flow = weighted_quantile(fl_vals, fl_w, 0.95) if fl_vals else None

                rows.append({
                    "z": z,
                    "layer": Lx,
                    "time_s": t,
                    "peak_flow": peak_flow,
                    "p95_flow": p95_flow,
                    "peak_speed": peak_speed,
                    "p95_speed": p95_speed,
                })

            rows.sort(key=lambda r: (float('inf') if r["z"] is None else r["z"]))
            return rows

        def _interp_by_z(rows, z_query):
            """Linear interpolation of a metric over Z.

            For per-layer series, we treat the value at each layer's Z and interpolate between adjacent Zs.
            If z_query is outside the known range, returns None.
            """
            pts = [(r["z"], r) for r in rows if r.get("z") is not None]
            if not pts:
                return None
            zs = [z for z, _ in pts]
            if z_query < zs[0] or z_query > zs[-1]:
                return None
            # exact match
            for z, rr in pts:
                if z == z_query:
                    return rr
            # find neighbors
            lo_i = 0
            hi_i = len(pts) - 1
            # binary search for insertion point
            import bisect
            idx = bisect.bisect_left(zs, z_query)
            if idx <= 0:
                return pts[0][1]
            if idx >= len(pts):
                return pts[-1][1]
            z0, r0 = pts[idx - 1]
            z1, r1 = pts[idx]
            if z1 == z0:
                return r1
            t = (z_query - z0) / (z1 - z0)
            out = {"z": z_query}
            # interpolate numeric keys
            for k in ("time_s", "peak_flow", "p95_flow", "peak_speed", "p95_speed"):
                v0 = r0.get(k)
                v1 = r1.get(k)
                if v0 is None or v1 is None:
                    out[k] = None
                else:
                    out[k] = (1 - t) * float(v0) + t * float(v1)
            # layer number: nearest by Z
            out["layer"] = r0.get("layer") if (z_query - z0) <= (z1 - z_query) else r1.get("layer")
            return out

        # Build Z-aligned comparison for the first compare run (overlay charts).
        first = compare_runs[0]
        A_label = run_label or "A"
        B_label = (first.get("label") or "B")

        a_cfg = config_info or {}
        b_cfg = first.get("config_info") or {}

        def _to_float(v):
            try:
                return float(v) if v is not None else None
            except Exception:
                return None

        flow_lim_a = _to_float(a_cfg.get("filament_max_volumetric_speed"))
        flow_lim_b = _to_float(b_cfg.get("filament_max_volumetric_speed"))
        sp_lim_a = _to_float(a_cfg.get("max_print_speed"))
        sp_lim_b = _to_float(b_cfg.get("max_print_speed"))

        A_rows = _layer_stats_series(moves, layer_z_map)
        B_rows = _layer_stats_series(first["moves"], first["layer_z_map"])

        zA = [r["z"] for r in A_rows if r.get("z") is not None]
        zB = [r["z"] for r in B_rows if r.get("z") is not None]
        z_common = sorted(set(zA) | set(zB))

        ws_cb.append([
            "Z_mm",
            f"{A_label}_layer", f"{A_label}_time_s", f"{A_label}_peak_flow", f"{A_label}_p95_flow", f"{A_label}_peak_speed",
            f"{B_label}_layer", f"{B_label}_time_s", f"{B_label}_peak_flow", f"{B_label}_p95_flow", f"{B_label}_peak_speed",
            f"{A_label}_limit_y", f"{B_label}_limit_y",
        ])

        # Fill data rows
        for z in z_common:
            a = _interp_by_z(A_rows, z)
            b = _interp_by_z(B_rows, z)
            ws_cb.append([
                z,
                a.get("layer") if a else None, a.get("time_s") if a else None, a.get("peak_flow") if a else None, a.get("p95_flow") if a else None, a.get("peak_speed") if a else None,
                b.get("layer") if b else None, b.get("time_s") if b else None, b.get("peak_flow") if b else None, b.get("p95_flow") if b else None, b.get("peak_speed") if b else None,
                None, None,
            ])

        data_end_row = ws_cb.max_row

        # Column widths
        set_basic_column_widths(
            ws_cb,
            {
                "A": 10,"B": 9,"C": 12,"D": 12,"E": 12,"F": 12,"G": 9,"H": 12,"I": 12,"J": 12,"K": 12,"L": 12,"M": 12,
            },
        )

        ws_cs = wb.create_sheet("Compare_Summary")
        header = ["Metric", (run_label or "A")]
        for r in compare_runs:
            header.extend([r.get("label") or Path(r.get("path","B")).stem, "Delta"])
        ws_cs.append(header)

        def _sum_or_none(d, key):
            vals = [v.get(key) for v in d.values() if v.get(key) is not None]
            return sum(vals) if vals else None

        def _metric_row(label, aval, bvals):
            row = [label, aval]
            for bv in bvals:
                row.append(bv)
                if aval is not None and bv is not None:
                    row.append(bv - aval)
                else:
                    row.append(None)
            return row

        def _layer_stats_from_moves(moves_x, layer_z_x):
            """Aggregate per-layer stats into a dict keyed by layer index.

            Values include:
              - time_s
              - peak_flow, p95_flow
              - peak_speed, p95_speed

            Percentiles are time-weighted.
            """
            by = defaultdict(list)
            for m in moves_x:
                by[m["layer"]].append(m)

            out = {}
            for Lx, msx in by.items():
                t = sum(m.get("time_s", 0.0) or 0.0 for m in msx)
                sp_vals = [m["speed_mm_s"] for m in msx if m.get("speed_mm_s") is not None and (m.get("dist_mm") or 0) > 0]
                sp_w = [m.get("time_s") or 0.0 for m in msx if m.get("speed_mm_s") is not None and (m.get("dist_mm") or 0) > 0]
                fl_vals = [m["flow_mm3_s"] for m in msx if (m.get("flow_mm3_s") or 0.0) > 0.0]
                fl_w = [m.get("time_s") or 0.0 for m in msx if (m.get("flow_mm3_s") or 0.0) > 0.0]

                out[Lx] = {
                    "z": layer_z_x.get(Lx, msx[-1].get("z")),
                    "time_s": t,
                    "peak_flow": max(fl_vals) if fl_vals else None,
                    "p95_flow": weighted_quantile(fl_vals, fl_w, 0.95) if fl_vals else None,
                    "peak_speed": max(sp_vals) if sp_vals else None,
                    "p95_speed": weighted_quantile(sp_vals, sp_w, 0.95) if sp_vals else None,
                }
            return out

        # Totals and maxima for each compare
        A_dict = _layer_stats_from_moves(moves, layer_z_map)
        total_a = _sum_or_none(A_dict, "time_s")
        b_dicts = [_layer_stats_from_moves(r["moves"], r["layer_z_map"]) for r in compare_runs]
        totals_b = [_sum_or_none(d, "time_s") for d in b_dicts]
        ws_cs.append(_metric_row("Total time (s)", total_a, totals_b))

        def _max_key(d, key):
            if not d:
                return None
            return max((v.get(key) or 0) for v in d.values())

        ws_cs.append(_metric_row("Max peak flow (mm³/s)", _max_key(A_dict, "peak_flow"), [_max_key(d, "peak_flow") for d in b_dicts]))
        ws_cs.append(_metric_row("Max P95 flow (mm³/s)", _max_key(A_dict, "p95_flow"), [_max_key(d, "p95_flow") for d in b_dicts]))
        ws_cs.append(_metric_row("Max peak speed (mm/s)", _max_key(A_dict, "peak_speed"), [_max_key(d, "peak_speed") for d in b_dicts]))
        ws_cs.append(_metric_row("Max P95 speed (mm/s)", _max_key(A_dict, "p95_speed"), [_max_key(d, "p95_speed") for d in b_dicts]))

        # Widths
        widths = {"A": 26, "B": 14}
        col = 3
        for _ in compare_runs:
            widths[get_column_letter(col)] = 14
            widths[get_column_letter(col + 1)] = 10
            col += 2
        set_basic_column_widths(ws_cs, widths)

        # Overlay charts on dashboard (comparison-focused)
        # Place them *below* the existing dashboard charts.
        compare_r1 = R9 + 36
        compare_r2 = compare_r1 + 18

        def _nice_major_unit(max_x, target_ticks=8):
            if max_x is None or max_x <= 0:
                return None
            raw = max_x / float(target_ticks)
            # round to 1/2/5 * 10^n
            exp = math.floor(math.log10(raw)) if raw > 0 else 0
            base = raw / (10 ** exp)
            if base <= 1:
                nice = 1
            elif base <= 2:
                nice = 2
            elif base <= 5:
                nice = 5
            else:
                nice = 10
            return nice * (10 ** exp)

        # Determine X axis range from Z (common axis)
        max_z = None
        try:
            vals = [r[0] for r in ws_cb.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True) if r[0] is not None]
            max_z = max(vals) if vals else None
        except Exception:
            max_z = None

        major = _nice_major_unit(max_z, target_ticks=7)

        # We keep helper points below the data table so charts can reference 2-point constant limit lines
        helper_row = (data_end_row or ws_cb.max_row) + 2

        def _add_compare_scatter(title, y_title, col_ay, col_by, anchor, limit_a=None, limit_b=None):
            nonlocal helper_row
            ch = ScatterChart()
            ch.title = title
            ch.y_axis.title = y_title
            ch.x_axis.title = "Z (mm)"
            _style_axis(ch.y_axis)
            _style_x_axis(ch.x_axis)

            # Force X/Y tick label font parity (Excel can otherwise enlarge X labels)
            try:
                ch.x_axis.txPr = ch.y_axis.txPr
            except Exception:
                pass

            if major is not None:
                ch.x_axis.majorUnit = major

            # Common X for both series
            x_ref = Reference(ws_cb, min_col=1, min_row=2, max_row=data_end_row)

            # A series
            y_a = Reference(ws_cb, min_col=col_ay, min_row=1, max_row=data_end_row)
            s_a = Series(y_a, x_ref, title_from_data=True)
            ch.series.append(s_a)

            # B series
            y_b = Reference(ws_cb, min_col=col_by, min_row=1, max_row=data_end_row)
            s_b = Series(y_b, x_ref, title_from_data=True)
            ch.series.append(s_b)

            # Optional per-run limit reference lines (2-point scatter)
            if limit_a is not None:
                ws_cb.cell(row=helper_row, column=1, value=0.0)
                ws_cb.cell(row=helper_row+1, column=1, value=max_z)
                ws_cb.cell(row=helper_row, column=12, value=float(limit_a))
                ws_cb.cell(row=helper_row+1, column=12, value=float(limit_a))
                y_lim_a = Reference(ws_cb, min_col=12, min_row=helper_row, max_row=helper_row+1)
                x_lim = Reference(ws_cb, min_col=1, min_row=helper_row, max_row=helper_row+1)
                s_la = Series(y_lim_a, x_lim, title=f"{A_label}_limit")
                ch.series.append(s_la)

            if limit_b is not None:
                ws_cb.cell(row=helper_row, column=13, value=float(limit_b))
                ws_cb.cell(row=helper_row+1, column=13, value=float(limit_b))
                y_lim_b = Reference(ws_cb, min_col=13, min_row=helper_row, max_row=helper_row+1)
                x_lim = Reference(ws_cb, min_col=1, min_row=helper_row, max_row=helper_row+1)
                s_lb = Series(y_lim_b, x_lim, title=f"{B_label}_limit")
                ch.series.append(s_lb)

            helper_row += 3

            # Layout
            ch.legend.position = "r"
            ch.height = CH_H
            ch.width = CH_W
            ws_dash.add_chart(ch, anchor)

        _add_compare_scatter("Compare: Layer Time (s)", "seconds", 3, 8, f"{LEFT}{compare_r1}")
        _add_compare_scatter("Compare: Peak Flow (mm³/s)", "mm³/s", 4, 9, f"{RIGHT}{compare_r1}", limit_a=flow_lim_a, limit_b=flow_lim_b)
        _add_compare_scatter("Compare: P95 Flow (mm³/s)", "mm³/s", 5, 10, f"{LEFT}{compare_r2}", limit_a=flow_lim_a, limit_b=flow_lim_b)
        _add_compare_scatter("Compare: Peak Speed (mm/s)", "mm/s", 6, 11, f"{RIGHT}{compare_r2}", limit_a=sp_lim_a, limit_b=sp_lim_b)

        # "Two x-axis scales" helper table (Z -> A layer, B layer) for major tick positions.
        try:
            if major is None:
                major_local = _nice_major_unit(max_z, target_ticks=7)
            else:
                major_local = major
            if major_local and max_z is not None:
                def _nearest_layer(rows, z_target):
                    best = None
                    for rr in rows:
                        z = rr.get("z") if isinstance(rr, dict) else None
                        layer = rr.get("layer") if isinstance(rr, dict) else None
                        if z is None:
                            continue
                        dz = abs(float(z) - float(z_target))
                        if best is None or dz < best[0]:
                            best = (dz, layer)
                    return best[1] if best else None

                ticks = []
                zt = 0.0
                while zt <= float(max_z) + 1e-9:
                    ticks.append(round(zt, 6))
                    zt += float(major_local)

                table_row = compare_r2 + int(CH_H * 2) + 3
                ws_dash[f"{LEFT}{table_row}"] = "Compare X-axis scales (by Z tick)"
                ws_dash[f"{LEFT}{table_row}"].alignment = Alignment(horizontal="left")
                ws_dash[f"{LEFT}{table_row+1}"] = "Z (mm)"
                ws_dash[f"{LEFT}{table_row+1}"].alignment = Alignment(horizontal="left")
                ws_dash[f"{LEFT}{table_row+1}"].fill = PatternFill("solid", fgColor="DDDDDD")
                ws_dash[f"{MIDDLE}{table_row+1}"] = "A layer"
                ws_dash[f"{MIDDLE}{table_row+1}"].fill = PatternFill("solid", fgColor="DDDDDD")
                ws_dash[f"{RIGHT}{table_row+1}"] = "B layer"
                ws_dash[f"{RIGHT}{table_row+1}"].fill = PatternFill("solid", fgColor="DDDDDD")

                r = table_row + 2
                for ztick in ticks[:20]:  # keep it compact
                    ws_dash[f"{LEFT}{r}"] = ztick
                    ws_dash[f"{MIDDLE}{r}"] = _nearest_layer(A_rows, ztick)
                    ws_dash[f"{RIGHT}{r}"] = _nearest_layer(B_rows, ztick)
                    r += 1
        except Exception:
            pass

    _status("Saving workbook")
    wb.save(out_path)


def _aggregate_layers_for_export(moves, layer_z_map, config_info=None):
    """Return list of dict rows matching the Layers sheet schema (subset).
    if RichText is None:
        return None

    This is used for CSV/JSON sidecars and tests. Keep it lightweight and stable.
    """
    by_layer = defaultdict(list)
    for m in moves:
        by_layer[m["layer"]].append(m)

    prev_z = None
    out = []
    for L in sorted(by_layer.keys()):
        ms = by_layer[L]
        z_val = layer_z_map.get(L, ms[-1].get("z"))
        layer_h = (z_val - prev_z) if (prev_z is not None and z_val is not None) else None
        if z_val is not None:
            prev_z = z_val

        t = sum(m.get("time_s") or 0.0 for m in ms)
        d = sum(m.get("dist_mm") or 0.0 for m in ms)
        e_pos = sum((m.get("de_mm") or 0.0) for m in ms if (m.get("de_mm") or 0.0) > 0.0)

        sp_vals = [m["speed_mm_s"] for m in ms if m.get("speed_mm_s") is not None and (m.get("dist_mm") or 0.0) > 0.0]
        sp_w = [m.get("time_s") or 0.0 for m in ms if m.get("speed_mm_s") is not None and (m.get("dist_mm") or 0.0) > 0.0]
        fl_vals = [m["flow_mm3_s"] for m in ms if (m.get("flow_mm3_s") or 0.0) > 0.0]
        fl_w = [m.get("time_s") or 0.0 for m in ms if (m.get("flow_mm3_s") or 0.0) > 0.0]

        peak_speed = max(sp_vals) if sp_vals else None
        p95_speed = weighted_quantile(sp_vals, sp_w, 0.95) if sp_vals else None
        p99_speed = weighted_quantile(sp_vals, sp_w, 0.99) if sp_vals else None
        peak_flow = max(fl_vals) if fl_vals else None
        p95_flow = weighted_quantile(fl_vals, fl_w, 0.95) if fl_vals else None
        p99_flow = weighted_quantile(fl_vals, fl_w, 0.99) if fl_vals else None

        flow_limit = (config_info or {}).get("filament_max_volumetric_speed")
        speed_limit = (config_info or {}).get("max_print_speed")
        flow_headroom = None
        speed_headroom = None
        try:
            if flow_limit is not None and p99_flow is not None:
                flow_headroom = float(flow_limit) - float(p99_flow)
        except Exception:
            pass
        try:
            if speed_limit is not None and p99_speed is not None:
                speed_headroom = float(speed_limit) - float(p99_speed)
        except Exception:
            pass

        travel_time = sum(m.get("time_s") or 0.0 for m in ms if (m.get("de_mm") or 0.0) == 0.0 and (m.get("dist_mm") or 0.0) > 0.0)
        travel_dist = sum(m.get("dist_mm") or 0.0 for m in ms if (m.get("de_mm") or 0.0) == 0.0 and (m.get("dist_mm") or 0.0) > 0.0)
        extrude_time = sum(m.get("time_s") or 0.0 for m in ms if (m.get("de_mm") or 0.0) > 0.0 and (m.get("time_s") or 0.0) > 0.0)
        retract_moves = [m for m in ms if (m.get("de_mm") or 0.0) < 0.0]
        retract_count = len(retract_moves)
        retract_mm = -sum(m.get("de_mm") or 0.0 for m in retract_moves)

        short_fast = 0
        for m in ms:
            if (m.get("de_mm") or 0.0) > 0.0 and (m.get("dist_mm") or 0.0) > 0.0:
                if (m.get("dist_mm") or 0.0) < 0.6 and (m.get("speed_mm_s") or 0.0) > 50.0:
                    short_fast += 1

        avg_speed = (d / t) if t > 0 else None
        avg_flow = (sum((m.get("flow_mm3_s") or 0.0) * (m.get("time_s") or 0.0) for m in ms) / t) if t > 0 else None

        out.append(
            {
                "layer": L,
                "z_mm": z_val,
                "layer_height_mm": layer_h,
                "time_s": t,
                "dist_mm": d,
                "extrusion_mm": e_pos,
                "avg_speed_mm_s": avg_speed,
                "avg_flow_mm3_s": avg_flow,
                "peak_speed_mm_s": peak_speed,
                "p95_speed_mm_s": p95_speed,
                "p99_speed_mm_s": p99_speed,
                "peak_flow_mm3_s": peak_flow,
                "p95_flow_mm3_s": p95_flow,
                "p99_flow_mm3_s": p99_flow,
                "flow_headroom_p99_mm3_s": flow_headroom,
                "speed_headroom_p99_mm_s": speed_headroom,
                "travel_time_s": travel_time,
                "travel_dist_mm": travel_dist,
                "extrude_time_s": extrude_time,
                "retract_count": retract_count,
                "retract_mm": retract_mm,
                "dynamics_score": short_fast,
            }
        )

    return out


def build_json_summary(moves, layer_z_map, config_info=None):
    """Build a small, regression-friendly summary object."""
    layers = _aggregate_layers_for_export(moves, layer_z_map, config_info=config_info)
    total_time_s = sum(r["time_s"] for r in layers)
    total_travel_time_s = sum(r["travel_time_s"] for r in layers)
    total_extrude_time_s = sum(r["extrude_time_s"] for r in layers)
    total_retracts = sum(r["retract_count"] for r in layers)
    total_retract_mm = sum(r["retract_mm"] for r in layers)

    def _max_of(key):
        vals = [r.get(key) for r in layers if r.get(key) is not None]
        return max(vals) if vals else None

    return {
        "layers": len(layers),
        "total_time_s": total_time_s,
        "total_travel_time_s": total_travel_time_s,
        "total_extrude_time_s": total_extrude_time_s,
        "total_retract_count": total_retracts,
        "total_retract_mm": total_retract_mm,
        "max_peak_speed_mm_s": _max_of("peak_speed_mm_s"),
        "max_p95_speed_mm_s": _max_of("p95_speed_mm_s"),
        "max_p99_speed_mm_s": _max_of("p99_speed_mm_s"),
        "max_peak_flow_mm3_s": _max_of("peak_flow_mm3_s"),
        "max_p95_flow_mm3_s": _max_of("p95_flow_mm3_s"),
        "max_p99_flow_mm3_s": _max_of("p99_flow_mm3_s"),
        "config": {
            "max_print_speed": (config_info or {}).get("max_print_speed"),
            "filament_max_volumetric_speed": (config_info or {}).get("filament_max_volumetric_speed"),
        },
    }


def write_csv_exports(moves, layer_z_map, out_xlsx_path: str, config_info=None, top_n_segments: int = 200):
    """Write CSV exports next to the XLSX (layers + top segments + feature histogram)."""
    import csv

    out_xlsx = Path(out_xlsx_path)
    base = out_xlsx.with_suffix("")

    # Layers
    layers = _aggregate_layers_for_export(moves, layer_z_map, config_info=config_info)
    layers_path = base.with_name(base.name + "_layers.csv")
    if layers:
        with open(layers_path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=list(layers[0].keys()))
            w.writeheader()
            w.writerows(layers)

    # Top flow segments
    segs = [m for m in moves if (m.get("de_mm") or 0.0) > 0.0 and (m.get("time_s") or 0.0) > 0.0 and (m.get("flow_mm3_s") or 0.0) > 0.0]
    segs.sort(key=lambda m: (m.get("flow_mm3_s") or 0.0), reverse=True)
    seg_path = base.with_name(base.name + "_top_flow_segments.csv")
    with open(seg_path, "w", newline="", encoding="utf-8") as f:
        fieldnames = [
            "rank", "layer", "type", "z_mm", "x0", "y0", "x1", "y1", "dist_mm", "de_mm", "time_s", "speed_mm_s", "flow_mm3_s"
        ]
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for idx, m in enumerate(segs[:max(1, int(top_n_segments))], start=1):
            w.writerow(
                {
                    "rank": idx,
                    "layer": m.get("layer"),
                    "type": m.get("type"),
                    "z_mm": m.get("z"),
                    "x0": m.get("x0"),
                    "y0": m.get("y0"),
                    "x1": m.get("x1"),
                    "y1": m.get("y1"),
                    "dist_mm": m.get("dist_mm"),
                    "de_mm": m.get("de_mm"),
                    "time_s": m.get("time_s"),
                    "speed_mm_s": m.get("speed_mm_s"),
                    "flow_mm3_s": m.get("flow_mm3_s"),
                }
            )

    # Feature histogram
    fh_path = base.with_name(base.name + "_feature_flow_hist.csv")
    flow_values = [m.get("flow_mm3_s") for m in segs]
    if flow_values:
        lo = 0.0
        hi = max(flow_values)
        if (config_info or {}).get("filament_max_volumetric_speed") is not None:
            try:
                hi = max(hi, float((config_info or {}).get("filament_max_volumetric_speed")))
            except Exception:
                pass
        bins_spec = make_bins(lo, hi, 20)
        total_time = sum(m.get("time_s") or 0.0 for m in segs)
        by_type = defaultdict(list)
        for m in segs:
            by_type[m.get("type") or "UNKNOWN"].append(m)
        rows = []
        for t, ms in sorted(by_type.items(), key=lambda kv: sum(m.get("time_s") or 0.0 for m in kv[1]), reverse=True):
            for (b_lo, b_hi) in bins_spec:
                bt = 0.0
                for m in ms:
                    v = m.get("flow_mm3_s") or 0.0
                    if b_lo <= v < b_hi or (b_hi == bins_spec[-1][1] and b_lo <= v <= b_hi):
                        bt += m.get("time_s") or 0.0
                rows.append({"type": t, "bin_lo": b_lo, "bin_hi": b_hi, "time_s": bt, "time_pct": (bt / total_time) if total_time > 0 else None})
        with open(fh_path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=["type", "bin_lo", "bin_hi", "time_s", "time_pct"])
            w.writeheader()
            w.writerows(rows)

